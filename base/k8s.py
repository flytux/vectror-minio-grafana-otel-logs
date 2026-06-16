from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule


OUTPUT_FILE = "k8s_cluster_sizing_template_advanced.xlsx"


# ----------------------------
# 공통 스타일
# ----------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
GRAY_FILL = PatternFill("solid", fgColor="EDEDED")
HIGHLIGHT_FILL = PatternFill("solid", fgColor="BDD7EE")

RED_FILL = PatternFill("solid", fgColor="F4CCCC")
ORANGE_FILL = PatternFill("solid", fgColor="FCE5CD")
GREEN_FILL = PatternFill("solid", fgColor="D9EAD3")


def apply_header_style(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_all_borders(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.border = BORDER


def set_column_widths(ws, widths):
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def protect_sheet(ws, password="1234"):
    ws.protection.sheet = True
    ws.protection.password = password


# ----------------------------
# 입력 시트
# ----------------------------
def add_input_validations(ws):
    # 숫자만 허용, 0 이상
    dv_decimal_non_negative = DataValidation(
        type="decimal",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=False
    )
    dv_decimal_non_negative.error = "0 이상의 숫자만 입력하세요."
    dv_decimal_non_negative.errorTitle = "입력 오류"
    dv_decimal_non_negative.prompt = "숫자 값을 입력하세요."
    dv_decimal_non_negative.promptTitle = "입력 안내"
    ws.add_data_validation(dv_decimal_non_negative)

    # 비율 전용: 0 ~ 1.5
    dv_ratio = DataValidation(
        type="decimal",
        operator="between",
        formula1="0",
        formula2="1.5",
        allow_blank=False
    )
    dv_ratio.error = "비율 값은 0 ~ 1.5 범위로 입력하세요."
    dv_ratio.errorTitle = "비율 입력 오류"
    dv_ratio.prompt = "0 ~ 1.5 범위의 비율 값을 입력하세요."
    dv_ratio.promptTitle = "비율 입력 안내"
    ws.add_data_validation(dv_ratio)

    # 정수 전용: 0 이상
    dv_integer_non_negative = DataValidation(
        type="whole",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=False
    )
    dv_integer_non_negative.error = "0 이상의 정수만 입력하세요."
    dv_integer_non_negative.errorTitle = "정수 입력 오류"
    dv_integer_non_negative.prompt = "정수 값을 입력하세요."
    dv_integer_non_negative.promptTitle = "정수 입력 안내"
    ws.add_data_validation(dv_integer_non_negative)

    # 행별 적용 (값 열 = D)
    integer_rows = [10, 11, 12, 15, 20]  # 최소 Replica, 여유노드, AZ, max_pods, cp_node_count
    ratio_rows = [16, 17, 18, 19, 23, 24, 25]  # 비율/헤드룸
    decimal_rows = [2, 3, 4, 5, 6, 7, 8, 9, 13, 14, 21, 22]  # 나머지 수치

    for r in integer_rows:
        dv_integer_non_negative.add(f"D{r}")
    for r in ratio_rows:
        dv_ratio.add(f"D{r}")
    for r in decimal_rows:
        dv_decimal_non_negative.add(f"D{r}")


def format_input_sheet(ws):
    apply_header_style(ws, 1)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{ws.max_row}"

    center_cols = ["A", "E"]
    for row in range(2, ws.max_row + 1):
        ws[f"D{row}"].fill = YELLOW_FILL
        ws[f"D{row}"].number_format = '#,##0.00'
        for col in center_cols:
            ws[f"{col}{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"F{row}"].alignment = Alignment(wrap_text=True, vertical="top")

    # 일부 값 포맷 보정
    percent_rows = [16, 17, 18, 19, 23]
    for r in percent_rows:
        ws[f"D{r}"].number_format = "0.0%"

    set_column_widths(ws, {
        1: 14,
        2: 24,
        3: 26,
        4: 14,
        5: 12,
        6: 42,
    })

    tab = Table(displayName="tblInput", ref=f"A1:F{ws.max_row}")
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)

    add_input_validations(ws)
    apply_all_borders(ws)


def create_input_sheet(wb):
    ws = wb.active
    ws.title = "입력"
    ws.append(["구분", "항목", "입력키", "값", "단위", "설명"])

    rows = [
        ["트래픽", "목표 TPS", "target_tps", 1000, "req/s", "목표 초당 처리 요청 수"],
        ["트래픽", "피크 배수", "peak_multiplier", 1.5, "배", "평균 대비 최대 피크 배수"],
        ["트래픽", "운영 여유율", "ops_buffer", 1.2, "배", "안전 버퍼 반영"],
        ["성능", "평균 응답시간(초)", "avg_response_sec", 0.2, "sec", "평균 요청 처리 시간"],
        ["성능", "Pod당 처리 가능 TPS", "pod_tps_capacity", 120, "req/s", "애플리케이션 Pod 1개가 감당 가능한 TPS"],
        ["Pod 사양", "Pod CPU request", "pod_cpu_request", 0.5, "core", "Pod 1개 요청 CPU"],
        ["Pod 사양", "Pod CPU limit", "pod_cpu_limit", 1, "core", "Pod 1개 최대 CPU"],
        ["Pod 사양", "Pod Memory request", "pod_mem_request", 1024, "MiB", "Pod 1개 요청 메모리"],
        ["Pod 사양", "Pod Memory limit", "pod_mem_limit", 2048, "MiB", "Pod 1개 최대 메모리"],
        ["가용성", "최소 Replica 수", "min_replicas", 2, "개", "서비스 최소 복제 수"],
        ["가용성", "장애 대비 여유 노드 수", "failover_nodes", 1, "개", "N+1 또는 추가 여유 노드"],
        ["가용성", "AZ 수", "az_count", 3, "개", "가용영역 수"],
        ["노드 사양", "워커 노드 vCPU", "worker_vcpu", 8, "core", "워커 노드 1대 CPU"],
        ["노드 사양", "워커 노드 메모리", "worker_mem_mib", 32768, "MiB", "워커 노드 1대 메모리"],
        ["노드 사양", "노드당 최대 Pod 수", "max_pods_per_node", 30, "개", "워커 노드 1대에 배치 가능한 최대 Pod 수"],
        ["노드 사양", "시스템 예약 CPU 비율", "system_cpu_reserve_ratio", 0.1, "비율", "kube/system reserved CPU 비율"],
        ["노드 사양", "시스템 예약 메모리 비율", "system_mem_reserve_ratio", 0.1, "비율", "kube/system reserved 메모리 비율"],
        ["운영 정책", "목표 CPU 사용률", "target_cpu_utilization", 0.7, "비율", "HPA/운영 목표 CPU 사용률"],
        ["운영 정책", "목표 메모리 사용률", "target_mem_utilization", 0.75, "비율", "운영 목표 메모리 사용률"],
        ["운영 정책", "컨트롤플레인 노드 수", "cp_node_count", 3, "개", "권장 기본값"],
        ["운영 정책", "컨트롤플레인 노드 vCPU", "cp_node_vcpu", 4, "core", "컨트롤플레인 1대 CPU"],
        ["운영 정책", "컨트롤플레인 노드 메모리", "cp_node_mem_mib", 8192, "MiB", "컨트롤플레인 1대 메모리"],
        ["운영 정책", "노드 목표 최대 사용률", "node_target_utilization", 0.8, "비율", "노드 과밀 방지용 목표 사용률"],
        ["운영 정책", "Pod 헤드룸 비율", "pod_headroom_ratio", 1.15, "배", "Pod 산정 추가 버퍼"],
        ["운영 정책", "노드 헤드룸 비율", "node_headroom_ratio", 1.15, "배", "노드 산정 추가 버퍼"],
    ]
    for row in rows:
        ws.append(row)

    format_input_sheet(ws)


# ----------------------------
# 계산 시트
# ----------------------------
def format_calc_sheet(ws):
    apply_header_style(ws, 1)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:D{ws.max_row}"

    for row in range(2, ws.max_row + 1):
        ws[f"B{row}"].fill = GRAY_FILL
        ws[f"C{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"D{row}"].alignment = Alignment(wrap_text=True, vertical="top")

    set_column_widths(ws, {
        1: 28,
        2: 18,
        3: 10,
        4: 44,
    })

    percent_rows = [16, 17, 18, 19, 23, 57, 58]
    for r in percent_rows:
        ws[f"B{r}"].number_format = "0.0%"
    for r in range(2, ws.max_row + 1):
        if r not in percent_rows:
            ws[f"B{r}"].number_format = '#,##0.00'

    apply_all_borders(ws)
    protect_sheet(ws)


def create_calc_sheet(wb):
    ws = wb.create_sheet("계산")
    ws.append(["계산항목", "값", "단위", "설명"])

    rows = [
        ["목표 TPS", '=XLOOKUP("target_tps",tblInput[입력키],tblInput[값])', "req/s", "입력값 조회"],
        ["피크 배수", '=XLOOKUP("peak_multiplier",tblInput[입력키],tblInput[값])', "배", "입력값 조회"],
        ["운영 여유율", '=XLOOKUP("ops_buffer",tblInput[입력키],tblInput[값])', "배", "입력값 조회"],
        ["평균 응답시간", '=XLOOKUP("avg_response_sec",tblInput[입력키],tblInput[값])', "sec", "입력값 조회"],
        ["Pod당 처리 가능 TPS", '=XLOOKUP("pod_tps_capacity",tblInput[입력키],tblInput[값])', "req/s", "입력값 조회"],
        ["Pod CPU request", '=XLOOKUP("pod_cpu_request",tblInput[입력키],tblInput[값])', "core", "입력값 조회"],
        ["Pod CPU limit", '=XLOOKUP("pod_cpu_limit",tblInput[입력키],tblInput[값])', "core", "입력값 조회"],
        ["Pod Memory request", '=XLOOKUP("pod_mem_request",tblInput[입력키],tblInput[값])', "MiB", "입력값 조회"],
        ["Pod Memory limit", '=XLOOKUP("pod_mem_limit",tblInput[입력키],tblInput[값])', "MiB", "입력값 조회"],
        ["최소 Replica 수", '=XLOOKUP("min_replicas",tblInput[입력키],tblInput[값])', "개", "입력값 조회"],
        ["장애 대비 여유 노드 수", '=XLOOKUP("failover_nodes",tblInput[입력키],tblInput[값])', "개", "입력값 조회"],
        ["AZ 수", '=XLOOKUP("az_count",tblInput[입력키],tblInput[값])', "개", "입력값 조회"],
        ["워커 노드 vCPU", '=XLOOKUP("worker_vcpu",tblInput[입력키],tblInput[값])', "core", "입력값 조회"],
        ["워커 노드 메모리", '=XLOOKUP("worker_mem_mib",tblInput[입력키],tblInput[값])', "MiB", "입력값 조회"],
        ["노드당 최대 Pod 수", '=XLOOKUP("max_pods_per_node",tblInput[입력키],tblInput[값])', "개", "입력값 조회"],
        ["시스템 예약 CPU 비율", '=XLOOKUP("system_cpu_reserve_ratio",tblInput[입력키],tblInput[값])', "비율", "입력값 조회"],
        ["시스템 예약 메모리 비율", '=XLOOKUP("system_mem_reserve_ratio",tblInput[입력키],tblInput[값])', "비율", "입력값 조회"],
        ["목표 CPU 사용률", '=XLOOKUP("target_cpu_utilization",tblInput[입력키],tblInput[값])', "비율", "입력값 조회"],
        ["목표 메모리 사용률", '=XLOOKUP("target_mem_utilization",tblInput[입력키],tblInput[값])', "비율", "입력값 조회"],
        ["컨트롤플레인 노드 수", '=XLOOKUP("cp_node_count",tblInput[입력키],tblInput[값])', "개", "입력값 조회"],
        ["컨트롤플레인 노드 vCPU", '=XLOOKUP("cp_node_vcpu",tblInput[입력키],tblInput[값])', "core", "입력값 조회"],
        ["컨트롤플레인 노드 메모리", '=XLOOKUP("cp_node_mem_mib",tblInput[입력키],tblInput[값])', "MiB", "입력값 조회"],
        ["노드 목표 최대 사용률", '=XLOOKUP("node_target_utilization",tblInput[입력키],tblInput[값])', "비율", "노드 운영 목표"],
        ["Pod 헤드룸 비율", '=XLOOKUP("pod_headroom_ratio",tblInput[입력키],tblInput[값])', "배", "Pod 추가 버퍼"],
        ["노드 헤드룸 비율", '=XLOOKUP("node_headroom_ratio",tblInput[입력키],tblInput[값])', "배", "노드 추가 버퍼"],
        ["피크 TPS", "=B2*B3", "req/s", "목표 TPS x 피크 배수"],
        ["설계 TPS", "=B27*B4", "req/s", "피크 TPS x 운영 여유율"],
        ["필요 동시 처리 요청 수", "=B28*B5", "req", "설계 TPS x 평균 응답시간"],
        ["TPS 기준 필요 Pod 수", "=ROUNDUP(B28/B6,0)", "개", "설계 TPS / Pod당 처리 가능 TPS"],
        ["헤드룸 반영 Pod 수", "=ROUNDUP(B30*B24,0)", "개", "Pod 헤드룸 적용"],
        ["최소 Replica 반영 Pod 수", "=MAX(B31,B10)", "개", "최소 Replica 보장"],
        ["CPU 운영목표 반영 Pod 수", "=ROUNDUP(B32/B18,0)", "개", "목표 CPU 사용률 반영"],
        ["최종 필요 Pod 수", "=MAX(B32,B33)", "개", "최종 권장 Pod 수"],
        ["총 CPU request", "=B34*B6", "core", "최종 Pod 수 x Pod CPU request"],
        ["총 CPU limit", "=B34*B7", "core", "최종 Pod 수 x Pod CPU limit"],
        ["총 Memory request", "=B34*B8", "MiB", "최종 Pod 수 x Pod Memory request"],
        ["총 Memory limit", "=B34*B9", "MiB", "최종 Pod 수 x Pod Memory limit"],
        ["노드 가용 CPU", "=B13*(1-B16)*B23", "core", "시스템 예약 및 노드 목표 사용률 반영"],
        ["노드 가용 메모리", "=B14*(1-B17)*B23", "MiB", "시스템 예약 및 노드 목표 사용률 반영"],
        ["CPU 기준 필요 노드 수", "=ROUNDUP(B35/B39,0)", "개", "총 CPU request / 노드 가용 CPU"],
        ["메모리 기준 필요 노드 수", "=ROUNDUP(B37/B40,0)", "개", "총 Memory request / 노드 가용 메모리"],
        ["Pod 수 기준 필요 노드 수", "=ROUNDUP(B34/B15,0)", "개", "최종 필요 Pod 수 / 노드당 최대 Pod 수"],
        ["기본 필요 노드 수", "=MAX(B41,B42,B43)", "개", "세 기준 중 최대"],
        ["헤드룸 반영 노드 수", "=ROUNDUP(B44*B25,0)", "개", "노드 헤드룸 적용"],
        ["장애 여유 반영 노드 수", "=B45+B11", "개", "장애 대비 여유 노드 추가"],
        ["AZ 정렬 반영 노드 수", "=ROUNDUP(B46/B12,0)*B12", "개", "AZ 수 배수로 반올림"],
        ["클러스터 총 워커 CPU", "=B47*B13", "core", "워커 노드 총 CPU"],
        ["클러스터 총 워커 메모리", "=B47*B14", "MiB", "워커 노드 총 메모리"],
        ["컨트롤플레인 총 CPU", "=B20*B21", "core", "컨트롤플레인 총 CPU"],
        ["컨트롤플레인 총 메모리", "=B20*B22", "MiB", "컨트롤플레인 총 메모리"],
        ["전체 클러스터 총 CPU", "=B48+B50", "core", "워커+컨트롤플레인"],
        ["전체 클러스터 총 메모리", "=B49+B51", "MiB", "워커+컨트롤플레인"],
        ["권장 HPA 최소 Replica", "=B10", "개", "최소 Replica 기준"],
        ["권장 HPA 최대 Replica", "=ROUNDUP(B34*1.5,0)", "개", "최종 필요 Pod 수의 1.5배"],
        ["노드당 평균 Pod 수", "=ROUNDUP(B34/B47,0)", "개", "평균 Pod 배치 밀도"],
        ["노드 CPU 여유율", "=1-(B35/(B47*B13))", "비율", "전체 워커 CPU 대비 요청량"],
        ["노드 메모리 여유율", "=1-(B37/(B47*B14))", "비율", "전체 워커 메모리 대비 요청량"],
    ]
    for row in rows:
        ws.append(row)

    format_calc_sheet(ws)


# ----------------------------
# 결과 시트
# ----------------------------
def add_result_conditional_formatting(ws):
    # CPU / 메모리 여유율 (C22, C23)
    for cell in ["C21", "C22"]:
        ws.conditional_formatting.add(
            cell,
            CellIsRule(operator='lessThan', formula=['0.2'], fill=RED_FILL)
        )
        ws.conditional_formatting.add(
            cell,
            CellIsRule(operator='between', formula=['0.2', '0.35'], fill=ORANGE_FILL)
        )
        ws.conditional_formatting.add(
            cell,
            CellIsRule(operator='greaterThanOrEqual', formula=['0.35'], fill=GREEN_FILL)
        )

    # 노드당 평균 Pod 수 (C20) / max pods 30 기준으로 경고
    ws.conditional_formatting.add(
        "C20",
        CellIsRule(operator='greaterThan', formula=['27'], fill=RED_FILL)
    )
    ws.conditional_formatting.add(
        "C20",
        CellIsRule(operator='between', formula=['24', '27'], fill=ORANGE_FILL)
    )


def format_result_sheet(ws):
    apply_header_style(ws, 1)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{ws.max_row}"

    section_fills = {
        "요약": "D9EAF7",
        "애플리케이션": "E2F0D9",
        "워커 노드": "FFF2CC",
        "컨트롤플레인": "FCE4D6",
        "클러스터 합계": "D9D2E9",
        "오토스케일": "EAD1DC",
        "운영지표": "F4CCCC",
        "최종 권장안": "CFE2F3",
    }

    key_rows = [5, 8, 16, 17, 23]

    for row in range(2, ws.max_row + 1):
        section = ws[f"A{row}"].value
        if section in section_fills:
            fill = PatternFill("solid", fgColor=section_fills[section])
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = fill

        ws[f"D{row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"E{row}"].alignment = Alignment(wrap_text=True, vertical="top")

    for row in key_rows:
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = HIGHLIGHT_FILL
            ws.cell(row=row, column=col).font = Font(bold=True)

    set_column_widths(ws, {
        1: 16,
        2: 28,
        3: 24,
        4: 10,
        5: 30,
    })

    # 숫자/비율 서식
    for row in range(2, ws.max_row + 1):
        ws[f"C{row}"].number_format = '#,##0.00'
    for r in [21, 22]:
        ws[f"C{r}"].number_format = "0.0%"

    add_result_conditional_formatting(ws)
    apply_all_borders(ws)
    protect_sheet(ws)


def create_result_sheet(wb):
    ws = wb.create_sheet("결과")
    ws.append(["섹션", "항목", "값", "단위", "비고"])

    rows = [
        ["요약", "목표 TPS", "=계산!B2", "req/s", "사용자 입력값"],
        ["요약", "피크 TPS", "=계산!B26", "req/s", "피크 배수 반영"],
        ["요약", "설계 TPS", "=계산!B27", "req/s", "운영 여유율 반영"],
        ["요약", "평균 응답시간", "=계산!B4", "sec", "입력값"],
        ["애플리케이션", "최종 필요 Pod 수", "=계산!B34", "개", "권장 애플리케이션 Pod 수"],
        ["애플리케이션", "총 CPU request", "=계산!B35", "core", "애플리케이션 기준"],
        ["애플리케이션", "총 Memory request", "=ROUND(계산!B37/1024,1)", "GiB", "애플리케이션 기준"],
        ["워커 노드", "권장 워커 노드 수", "=계산!B47", "개", "AZ/장애여유/헤드룸 반영"],
        ["워커 노드", "워커 노드 1대 사양", '=계산!B13&" vCPU / "&ROUND(계산!B14/1024,1)&" GiB"', "", "표준 워커 노드"],
        ["워커 노드", "클러스터 총 워커 CPU", "=계산!B48", "core", "워커 노드 전체 합"],
        ["워커 노드", "클러스터 총 워커 메모리", "=ROUND(계산!B49/1024,1)", "GiB", "워커 노드 전체 합"],
        ["컨트롤플레인", "컨트롤플레인 노드 수", "=계산!B20", "개", "권장 기본 구성"],
        ["컨트롤플레인", "컨트롤플레인 1대 사양", '=계산!B21&" vCPU / "&ROUND(계산!B22/1024,1)&" GiB"', "", "기본 권장"],
        ["컨트롤플레인", "컨트롤플레인 총 CPU", "=계산!B50", "core", "전체 합"],
        ["컨트롤플레인", "컨트롤플레인 총 메모리", "=ROUND(계산!B51/1024,1)", "GiB", "전체 합"],
        ["클러스터 합계", "전체 클러스터 총 CPU", "=계산!B52", "core", "워커 + 컨트롤플레인"],
        ["클러스터 합계", "전체 클러스터 총 메모리", "=ROUND(계산!B53/1024,1)", "GiB", "워커 + 컨트롤플레인"],
        ["오토스케일", "권장 HPA 최소 Replica", "=계산!B54", "개", "최소값"],
        ["오토스케일", "권장 HPA 최대 Replica", "=계산!B55", "개", "최대값"],
        ["운영지표", "노드당 평균 Pod 수", "=계산!B56", "개", "배치 밀도"],
        ["운영지표", "노드 CPU 여유율", "=계산!B57", "비율", "낮을수록 빡빡함"],
        ["운영지표", "노드 메모리 여유율", "=계산!B58", "비율", "낮을수록 빡빡함"],
        ["최종 권장안", "권장 구성 요약", '="워커 노드 "&계산!B47&"대 / 컨트롤플레인 "&계산!B20&"대 / 총 Pod "&계산!B34&"개"', "", "최종 요약"],
    ]
    for row in rows:
        ws.append(row)

    format_result_sheet(ws)


# ----------------------------
# 메인
# ----------------------------
def main():
    wb = Workbook()
    create_input_sheet(wb)
    create_calc_sheet(wb)
    create_result_sheet(wb)

    # 활성 시트 결과로 설정
    wb.active = wb["결과"]

    wb.save(OUTPUT_FILE)
    print(f"생성 완료: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
